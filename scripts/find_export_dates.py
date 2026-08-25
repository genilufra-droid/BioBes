from pathlib import Path
from openpyxl import load_workbook

path = Path('/home/ubuntu/upload/2026FATURAT,LIKUJDIMETEFERMEREVE(1)(4).xlsx')
wb = load_workbook(path, data_only=False, read_only=True)
ws = wb['EKSPORTI']
for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
    if len(row) > 2 and str(row[1]).strip() in {'140', '195', '540'}:
        print(row_idx, list(row[:18]))
