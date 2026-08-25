from pathlib import Path
from openpyxl import load_workbook

path = Path('/home/ubuntu/upload/2026FATURAT,LIKUJDIMETEFERMEREVE(1)(4).xlsx')
wb = load_workbook(path, data_only=False, read_only=True)
for ws in wb.worksheets:
    hits = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
        values = [cell.value for cell in row]
        text = ' | '.join('' if value is None else str(value) for value in values)
        if '540' in text or '2030' in text:
            hits.append((row_idx, values[:18]))
    if hits:
        print(f'[{ws.title}]')
        for row_idx, values in hits[:20]:
            print(row_idx, values)
