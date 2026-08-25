from pathlib import Path
from openpyxl import load_workbook

path = Path('/home/ubuntu/upload/2026FATURAT,LIKUJDIMETEFERMEREVE(1)(4).xlsx')
wb = load_workbook(path, data_only=False, read_only=True)
print('SHEETS:', wb.sheetnames)
for ws in wb.worksheets:
    print(f'\nSHEET {ws.title!r} max_row={ws.max_row} max_column={ws.max_column}')
    for row_no in range(1, min(ws.max_row, 45) + 1):
        values = [ws.cell(row_no, col).value for col in range(1, min(ws.max_column, 12) + 1)]
        if row_no <= 8 or row_no in (20,21,22,23,24,25,26,34,35,36,37,38,39,40):
            print(row_no, values)
