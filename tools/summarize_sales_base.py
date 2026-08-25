from pathlib import Path
from collections import Counter, defaultdict
from openpyxl import load_workbook
import json

source = Path('/home/ubuntu/upload/2026FATURAT,LIKUJDIMETEFERMEREVE(1)(4).xlsx')
output = Path('/home/ubuntu/sistemi-genit-cloud/sales-base-summary.json')
wb = load_workbook(source, data_only=True, read_only=True)
ws = wb['FATURAT 2026']
rows = [list(r) for r in ws.iter_rows(values_only=True)]
# Locate the real header row containing Data/Date and use only rows with a valid date thereafter.
header_index = next(i for i, row in enumerate(rows) if any(str(v).strip() == 'Data / Date' for v in row if v is not None))
header = rows[header_index]
records = []
for row in rows[header_index + 1:]:
    if not row or not isinstance(row[0], __import__('datetime').datetime):
        continue
    records.append(row)
def val(row, i):
    return row[i] if i < len(row) else None

def num(value):
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(' ', '').replace(',', ''))
    except Exception:
        return 0.0

def text(value):
    return str(value).strip() if value is not None else ''
invoice_groups = defaultdict(lambda: {'rows': 0, 'total': 0.0, 'customers': set(), 'dates': set()})
for r in records:
    invoice = text(val(r, 1))
    g = invoice_groups[invoice]
    g['rows'] += 1
    g['total'] += num(val(r, 8))
    g['customers'].add(text(val(r, 3)))
    g['dates'].add(text(val(r, 0).date()))
for g in invoice_groups.values():
    g['customers'] = sorted(x for x in g['customers'] if x)
    g['dates'] = sorted(x for x in g['dates'] if x)
summary = {
    'sheet': ws.title,
    'header_index_zero_based': header_index,
    'headers_first_21': [text(x) for x in header[:21]],
    'record_count': len(records),
    'date_min': str(min(val(r, 0) for r in records).date()) if records else None,
    'date_max': str(max(val(r, 0) for r in records).date()) if records else None,
    'invoice_count': len([k for k in invoice_groups if k]),
    'unique_invoice_numbers': sorted(k for k in invoice_groups if k)[:30],
    'invoice_samples': {k: invoice_groups[k] for k in sorted(invoice_groups)[:10] if k},
    'customer_column_candidates': {
      'seller_or_customer': sorted(Counter(text(val(r, 3)) for r in records).items(), key=lambda x: (-x[1], x[0]))[:20],
      'company_column': sorted(Counter(text(val(r, 11)) for r in records).items(), key=lambda x: (-x[1], x[0]))[:20],
    },
    'product_names': sorted(Counter(text(val(r, 5)) for r in records).items(), key=lambda x: (-x[1], x[0]))[:30],
    'currency_like_values': sorted(Counter(text(val(r, 17)) for r in records).items(), key=lambda x: (-x[1], x[0]))[:20],
    'totals': {
      'net_value': round(sum(num(val(r, 8)) for r in records), 2),
      'vat': round(sum(num(val(r, 9)) for r in records), 2),
      'gross_value': round(sum(num(val(r, 10)) for r in records), 2),
      'quantity_kg': round(sum(num(val(r, 6)) for r in records), 2),
    },
}
output.write_text(json.dumps(summary, ensure_ascii=False, indent=2, default=str), encoding='utf-8')
print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))
wb.close()
