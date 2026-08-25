from pathlib import Path
from openpyxl import load_workbook
import json

source = Path('/home/ubuntu/upload/2026FATURAT,LIKUJDIMETEFERMEREVE(1)(4).xlsx')
output = Path('/home/ubuntu/sistemi-genit-cloud/sales-sheets-detail.json')
wb = load_workbook(source, data_only=True, read_only=True)
selected = ['FATURAT 2026', 'PERMBLEDHJA E LIKUJDIMEVE', 'LIKUJDIMET NE BANKE', 'EKSPORTI', 'SHITJET B V NE LEKE & EURO', 'LISTE KLIENTESH']
result = {}
for name in selected:
    ws = wb[name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        vals = list(row)
        if any(v is not None and str(v).strip() for v in vals):
            rows.append(vals)
        if len(rows) >= 12:
            break
    result[name] = {'rows': rows, 'max_row': ws.max_row, 'max_column': ws.max_column}
output.write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str), encoding='utf-8')
print(output)
for name, data in result.items():
    print(f'\n## {name} ({data["max_row"]}x{data["max_column"]})')
    for row in data['rows'][:8]:
        print(' | '.join('' if v is None else str(v) for v in row[:24]))
wb.close()
