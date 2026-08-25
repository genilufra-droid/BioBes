from pathlib import Path
from openpyxl import load_workbook
import json

source = Path('/home/ubuntu/upload/2026FATURAT,LIKUJDIMETEFERMEREVE(1)(4).xlsx')
output = Path('/home/ubuntu/sistemi-genit-cloud/sales-workbook-analysis.json')
wb = load_workbook(source, data_only=False, read_only=True)
result = {'file': source.name, 'sheets': []}
for ws in wb.worksheets:
    rows = list(ws.iter_rows(values_only=False))
    nonempty = []
    for row in rows:
        values = [cell.value for cell in row]
        if any(value is not None and str(value).strip() != '' for value in values):
            nonempty.append(values)
    headers = nonempty[0] if nonempty else []
    samples = nonempty[1:4] if len(nonempty) > 1 else []
    formula_count = sum(1 for row in rows for cell in row if isinstance(cell.value, str) and cell.value.startswith('='))
    result['sheets'].append({
        'title': ws.title,
        'reported_dimensions': [ws.max_row, ws.max_column],
        'nonempty_rows': len(nonempty),
        'headers': headers,
        'samples': samples,
        'formula_count': formula_count,
    })
output.write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str), encoding='utf-8')
print(output)
for sheet in result['sheets']:
    print(f"{sheet['title']} | rows={sheet['nonempty_rows']} | cols={len(sheet['headers'])} | formulas={sheet['formula_count']}")
    print('  HEADERS:', ' | '.join(str(x) for x in sheet['headers'][:24]))
    for sample in sheet['samples'][:1]:
        print('  SAMPLE:', ' | '.join(str(x) for x in sample[:12]))
wb.close()
